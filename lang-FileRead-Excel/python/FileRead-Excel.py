import os

# it may need to install openpyxl
# >pip install openpyxl
from openpyxl import load_workbook

filename = os.path.abspath('.') + r"\..\DataFiles\data.xlsx"

def func_read_xlsx_by_cell_name(_file, _sheet, _cell):
    xlsx_data = load_workbook(_file, data_only=True)
    sheet1 = xlsx_data[_sheet]
    value = sheet1[_cell].value
    xlsx_data.close()
    return value
    
def func_read_xlsx_by_cell_position(_file, _sheet, _row, _col):
    xlsx_data = load_workbook(_file, data_only=True)
    sheet1 = xlsx_data[_sheet]
    value = sheet1.cell(_row, _col).value
    xlsx_data.close()
    return value


if __name__ == '__main__':

    xlsx_data = load_workbook(filename, data_only=True)
    sheet1 = xlsx_data['Sheet1']

    # Cell Data
    print('A1= ', sheet1['A1'].value)
    print('Cell(1,1)= ', sheet1.cell(1,1).value)

    # Multi-Cells 
    cells = sheet1['A1':'B2']
    for irow, row in enumerate(cells):
        for icol, col in enumerate(row):
            print('[', irow, ',', icol,']=', col.value)

    # Rows
    for row in sheet1.rows:
        print(row)
    
    # cols
    for col in sheet1.columns:
        print(col)
    
    xlsx_data.close()
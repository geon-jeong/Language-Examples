import os

# it may need to install openpyxl
# >pip install openpyxl
from openpyxl import load_workbook, Workbook


rfilename = os.path.abspath('.') + r"\..\DataFiles\data.xlsx"
wfilename = os.path.abspath('.') + r"\data.xlsx"

def func_read_xlsx_by_cell_name(_file, _sheet, _cell):
    xlsx_data = load_workbook(_file, data_only=True)
    sheet1 = xlsx_data[_sheet]
    value = sheet1[_cell].value
    xlsx_data.close()
    return value

    
def func_read_xlsx_by_cell_position(_file, _sheet, _row, _col):
    xlsx_data = load_workbook(_file, data_only=True)
    sheet1 = xlsx_data[_sheet]
    value = sheet1.cell(_row, _col).value
    xlsx_data.close()
    return value

    
if __name__ == '__main__':
    
    # Create new Excel Workbook
    w_workbook = Workbook()
    # Create new sheet. Note that the default 'Sheet' is already created. 
    w_sheet1 = w_workbook.create_sheet('Sheet1_ws')

    # write Cell 
    w_sheet1['A1'] = func_read_xlsx_by_cell_name(rfilename, 'Sheet1', 'A1')
    w_sheet1.cell(row=1, column=2, 
            value =func_read_xlsx_by_cell_position(rfilename, 'Sheet1', 2, 1))

    # write a row data
    w_sheet1.append([3,4,5])

    w_workbook.save(wfilename)
    w_workbook.close()


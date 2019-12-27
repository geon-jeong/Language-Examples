import os

# it may need to install openpyxl
# >pip install openpyxl
from openpyxl import load_workbook


filename = os.path.abspath('.') + r"\..\DataFiles\data.xlsx"

xlsx_data = load_workbook(filename, data_only=True)
sheet1 = xlsx_data['Sheet1']

# Cell Data
print('A1= ', sheet1['A1'].value)
print('Cell(1,1)= ', sheet1.cell(1,1).value)

# Multi-Cells 
cells = sheet1['A1':'B2']
for irow, row in enumerate(cells):
    for icol, col in enumerate(row):
        print('[', irow, ',', icol,']=', col.value)

# Rows
for row in sheet1.rows:
    print(row)
    
# cols
for col in sheet1.columns:
    print(col)